from tkinter import *
import tkinter.ttk as ttk
import tkinter.messagebox as msgbox
import tkinter.filedialog as fd
import serial
import serial.tools.list_ports
from openpyxl import Workbook, load_workbook
from PIL import Image, ImageTk
from gtts import gTTS
from time import sleep
arduino = None
main_frame = None
sdpath = None

def about():
    msgbox.showinfo(title="About",message="ShopSense-Connect\nThis software is used to manage the data on the ShopSense Device.")

def readSerial():
    data = str(arduino.readline()).strip("'b ")[:-4]
    while(data == ""):
        root.update()
        data = str(arduino.readline()).strip("'b ")[:-4]
    return data

def upload():
    path = fd.askopenfilename(filetypes=(
        ("Excel files", "*.xlsx"),
        ("All Files", "*.*")))
    if(path == ''):
        msgbox.showinfo(title="Failed", message="Please select a File.")
        return
    
    main_frame.destroy()

    upload_frame = LabelFrame(root, text="Uploading From "+path)

    upload_lbl = Label(upload_frame, text="Tap the RIFD Card: ---")
    upload_lbl.grid(row=0,column=0,padx=0,pady=0)

    img = Image.open("RIFD.png")
    img = img.resize((250, 250))
    img = ImageTk.PhotoImage(img)
    panel = Label(upload_frame, image=img)
    panel.image = img
    panel.grid(row=1,column=0)

    upload_frame.pack()

    wb = load_workbook(path)
    ws = wb.active
    for row in ws:
        if(row[0].value == 'SNo'):
            continue
        while True:
            sleep(1)
            arduino.write(b'upload')
            index = (int(row[0].value)-1)*4
            print(index)
            sleep(1)
            arduino.write([index])
            sleep(1)
            upload_lbl.config(text="Tap the RIFD Card: "+row[1].value)
            msgbox.showinfo(title="info", message="Tap the Card and press OK.")
            if(str(arduino.readall()).strip("'b ")[:-4] == "done"):
                print("done")
                break
            msgbox.showerror(title="Not Detected", message="Press OK and Tap the Card again.")
        tts = gTTS(str(row[1].value)+", Price "+str(row[2].value))
        if(len(str(row[0].value)) == 1): file = "0"+str(row[0].value)
        elif(len(str(row[0].value)) == 2): file = str(row[0].value)
        tts.save(sdpath+'/'+file+".mp3")
    
    msgbox.showinfo(title="Success", message="Successfully uploaded.")
    exit()    

def GenXl():
    path = fd.asksaveasfilename(filetypes=(
        ("Excel files", "*.xlsx"),
        ("All Files", "*.*")))
    if(path == ''):
        msgbox.showinfo(title="Failed", message="The Excel file is not successfully generated.\n File not selected.")
        return
    wb = Workbook()
    ws = wb.active
    ws.append(['SNo','Product Name','Price'])
    wb.save(path+".xlsx")
    msgbox.showinfo(title="Success", message="The Excel file is successfully generated.")
    

def sdpath():
    path = fd.askdirectory()
    path_lbl.config(text = path)
    
def SSConnect():
    global sdpath
    sdpath = path_lbl['text']
    if(sdpath == '-- Please Select the Path --'):
        msgbox.showerror(title="Failed", message="The ShopSense Module is not successfully connected.\nPlease select the path to SDCard.")
        return
    cp = comports[portnames.index(port_clicked.get())]
    global arduino
    arduino = serial.Serial(port=cp,   baudrate=9600, timeout=.1)
    data = readSerial()
    if(data == "ShopSense"):
        msgbox.showinfo(title="Success", message="The ShopSense Module is successfully connected.")
    else:
        msgbox.showerror(title="Failed", message="The ShopSense Module is not successfully connected.\nPlease select correct COM Port.")
        arduino.close()
        return
    welcome_frame.destroy()

    #filemenu.entryconfig(1,state='normal')
    filemenu.entryconfig(2,state='normal')
    global main_frame
    main_frame = ttk.LabelFrame(root, text = 'ShopSense on '+cp)
    
    img = Image.open("Logo.png")
    img = img.resize((250, 250))
    img = ImageTk.PhotoImage(img)
    panel = Label(main_frame, image=img)
    panel.image = img
    panel.grid(row=0,column=0)

    new_btn = ttk.Button(main_frame, text = "Upload new data", command = upload)
    new_btn.grid(row = 1,column = 0, padx = 10, pady = 10)
    
    load_btn = ttk.Button(main_frame, text = "Load existing data", command = GenXl, state="disabled")
    load_btn.grid(row = 2,column = 0, padx = 10, pady = 10)
    
    main_frame.pack(fill = 'both', padx = 10, pady = 10)

root = Tk()
root.title("ShopSense-Connect")

menubar = Menu(root)

filemenu = Menu(menubar, tearoff=0)
filemenu.add_command(label="New", command=GenXl)
filemenu.add_command(label="Load", command=GenXl, state='disabled')
filemenu.add_command(label="Upload", command=upload, state='disabled')
filemenu.add_separator()
filemenu.add_command(label="Exit", command=root.destroy)
menubar.add_cascade(label="File", menu=filemenu)

helpmenu = Menu(menubar, tearoff=0)
helpmenu.add_command(label="Help Manual", command=GenXl, state='disabled')
helpmenu.add_command(label="About", command=about)
menubar.add_cascade(label="Help", menu=helpmenu)

welcome_frame = ttk.LabelFrame(root, text = 'Welcome to ShopSense-Connect')

img = Image.open("Logo.png")
img = img.resize((250, 250))
img = ImageTk.PhotoImage(img)
panel = Label(welcome_frame, image=img)
panel.image = img
panel.grid(row=0,column=1)

port_lbl = ttk.Label(welcome_frame, text = "Port: ")
port_lbl.grid(row = 1, column = 0, padx = 10, pady = 10)

ports = serial.tools.list_ports.comports()
comports = []
portnames = []
for port, desc, hwid in sorted(ports):
    comports.append(port)
    portnames.append(desc)
port_clicked = StringVar()
port_clicked.set(portnames[0])

drop = ttk.OptionMenu( welcome_frame , port_clicked , *portnames) 
drop.grid(row = 1, column = 1, padx = 10, pady = 10)

sdpath_lbl = ttk.Label(welcome_frame, text = "Sdcard Path: ")
sdpath_lbl.grid(row = 2, column = 0, padx = 10, pady = 10)

path_lbl = ttk.Label(welcome_frame, text = "-- Please Select the Path --")
path_lbl.grid(row = 2, column = 1, padx = 10, pady = 10)

path_btn = ttk.Button(welcome_frame, text = "Select", command = sdpath)
path_btn.grid(row = 2,column = 2, padx = 10, pady = 10)

connect_btn = ttk.Button(welcome_frame, text = "Connect", command = SSConnect)
connect_btn.grid(row = 3,column = 1, padx = 10, pady = 10)


welcome_frame.pack(padx = 5 , pady = 5)
root.resizable(False, False)
root.config(menu=menubar)
root.iconphoto(False, PhotoImage(file="Logo.png"))
root.mainloop()